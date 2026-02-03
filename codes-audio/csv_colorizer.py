import os
import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, PieChart, Reference
import warnings
warnings.filterwarnings('ignore')

# ✅ Chemin du fichier CSV
csv_file = r"C:\Users\ADmiN\Desktop\video_project\faces_annotations_ultra.csv"

class FacialDataCleaner:
    def __init__(self, csv_path):
        self.csv_path = csv_path
        self.df = None
        self.cleaned_df = None
        
    def load_data(self):
        """Chargement des données depuis CSV"""
        try:
            self.df = pd.read_csv(self.csv_path)
            print(f"✅ {len(self.df)} enregistrements chargés du fichier")
            return True
        except Exception as e:
            print(f"❌ Erreur lors du chargement du fichier: {e}")
            return False
    
    def clean_data(self):
        """Nettoyage des données"""
        if self.df is None:
            return False
            
        self.cleaned_df = self.df.copy()
        
        # 🧹 Nettoyage des valeurs manquantes
        numeric_cols = ['face_confidence', 'age', 'gender_confidence', 
                       'emotion_confidence', 'race_confidence']
        
        for col in numeric_cols:
            if col in self.cleaned_df.columns:
                self.cleaned_df[col] = pd.to_numeric(self.cleaned_df[col], errors='coerce')
                # Arrondir les nombres à 3 décimales
                self.cleaned_df[col] = self.cleaned_df[col].round(3)
        
        # 🔧 Nettoyage des textes
        text_cols = ['video', 'image', 'gender', 'emotion', 'race']
        for col in text_cols:
            if col in self.cleaned_df.columns:
                self.cleaned_df[col] = self.cleaned_df[col].astype(str)
                self.cleaned_df[col] = self.cleaned_df[col].str.strip()
                self.cleaned_df[col] = self.cleaned_df[col].str.title()
        
        # 📊 Ajout d'une colonne de qualité globale
        if 'face_confidence' in self.cleaned_df.columns:
            self.cleaned_df['quality_score'] = self.calculate_quality_score()
            
        # 🎯 Ajout d'une catégorie d'âge
        if 'age' in self.cleaned_df.columns:
            self.cleaned_df['age_category'] = self.categorize_age()
            
        # ✨ Organisation des colonnes
        column_order = ['video', 'image', 'face_detected', 'face_confidence', 
                       'quality_score', 'age', 'age_category', 'gender', 
                       'gender_confidence', 'emotion', 'emotion_confidence', 
                       'race', 'race_confidence']
        
        # Réorganisation des colonnes existantes uniquement
        existing_cols = [col for col in column_order if col in self.cleaned_df.columns]
        remaining_cols = [col for col in self.cleaned_df.columns if col not in existing_cols]
        final_order = existing_cols + remaining_cols
        
        self.cleaned_df = self.cleaned_df[final_order]
        
        print(f"🧹 Données nettoyées avec succès!")
        return True
    
    def calculate_quality_score(self):
        """Calcul du score de qualité global du visage"""
        quality = []
        for _, row in self.cleaned_df.iterrows():
            if not row.get('face_detected', False):
                quality.append(0.0)
                continue
                
            score = 0.0
            weight_sum = 0.0
            
            # Poids de la confiance dans la détection du visage
            if pd.notna(row.get('face_confidence')):
                score += row['face_confidence'] * 0.4
                weight_sum += 0.4
                
            # Poids de la confiance du genre
            if pd.notna(row.get('gender_confidence')):
                score += row['gender_confidence'] * 0.2
                weight_sum += 0.2
                
            # Poids de la confiance des émotions
            if pd.notna(row.get('emotion_confidence')):
                score += row['emotion_confidence'] * 0.2
                weight_sum += 0.2
                
            # Poids de la confiance de l'origine
            if pd.notna(row.get('race_confidence')):
                score += row['race_confidence'] * 0.2
                weight_sum += 0.2
            
            if weight_sum > 0:
                quality.append(round(score / weight_sum, 3))
            else:
                quality.append(0.0)
                
        return quality
    
    def categorize_age(self):
        """Classification des âges en catégories"""
        categories = []
        for age in self.cleaned_df['age']:
            if pd.isna(age):
                categories.append('Unknown')
            elif age < 13:
                categories.append('Child')
            elif age < 20:
                categories.append('Teen')
            elif age < 35:
                categories.append('Young Adult')
            elif age < 55:
                categories.append('Adult')
            else:
                categories.append('Senior')
        return categories

class ExcelStyler:
    def __init__(self):
        # 🎨 Palette de couleurs améliorée
        self.colors = {
            'header': 'FF4A90E2',      # Bleu profond
            'file_info': 'FFE3F2FD',   # Bleu très clair
            'detection': 'FFF3E5F5',   # Violet clair
            'demographics': 'FFF1F8E9', # Vert clair
            'emotions': 'FFFFF3E0',    # Orange clair
            'confidence': 'FFFCE4EC',  # Rose clair
            'quality': 'FFE8F5E8',     # Vert menthe
            'high_quality': 'FF4CAF50', # Vert
            'medium_quality': 'FFFF9800', # Orange
            'low_quality': 'FFF44336',  # Rouge
            'detected_face': 'FFC8E6C9', # Vert clair
            'no_face': 'FFFFCDD2'      # Rouge clair
        }
        
    def get_column_category(self, col_name):
        """Détermination de la catégorie de la colonne"""
        col_lower = col_name.lower()
        if 'video' in col_lower or 'image' in col_lower:
            return 'file_info'
        elif 'face_detected' in col_lower or 'face_confidence' in col_lower:
            return 'detection'
        elif 'age' in col_lower or 'gender' in col_lower:
            return 'demographics'
        elif 'emotion' in col_lower:
            return 'emotions'
        elif 'confidence' in col_lower:
            return 'confidence'
        elif 'quality' in col_lower:
            return 'quality'
        else:
            return 'file_info'
    
    def apply_styles(self, ws, df):
        """Application de la mise en forme sur la feuille de calcul"""
        # 🎯 Formatage de l'en-tête
        for col_idx, col_name in enumerate(df.columns, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = PatternFill(start_color=self.colors['header'], 
                                  end_color=self.colors['header'], 
                                  fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF", size=11)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        
        # 🎨 Coloration des données
        for row_idx in range(2, ws.max_row + 1):
            for col_idx, col_name in enumerate(df.columns, 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                
                # Détermination de la couleur de fond selon le type de colonne
                category = self.get_column_category(col_name)
                fill_color = self.colors.get(category, 'FFFFFF')
                
                # Coloration spéciale pour les valeurs booléennes
                if col_name.lower() == 'face_detected':
                    if str(cell.value).lower() == 'true':
                        cell.value = "✅ Yes"
                        fill_color = self.colors['detected_face']
                    else:
                        cell.value = "❌ No"
                        fill_color = self.colors['no_face']
                
                # Coloration du score de qualité
                elif col_name.lower() == 'quality_score' and cell.value:
                    try:
                        score = float(cell.value)
                        if score >= 0.7:
                            fill_color = self.colors['high_quality']
                            cell.font = Font(color="FFFFFF", bold=True)
                        elif score >= 0.4:
                            fill_color = self.colors['medium_quality']
                        else:
                            fill_color = self.colors['low_quality']
                            cell.font = Font(color="FFFFFF")
                    except:
                        pass
                
                cell.fill = PatternFill(start_color=fill_color, 
                                      end_color=fill_color, 
                                      fill_type="solid")
                
                # Formatage du texte
                if not cell.font.color:
                    cell.font = Font(size=10)
                
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
        
        # 📏 Ajustement de la largeur des colonnes
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)  # Maximum 50
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # 🔒 Gel de la première ligne
        ws.freeze_panes = "A2"

def create_statistics_sheet(wb, df):
    """Création de la feuille de statistiques"""
    ws = wb.create_sheet("📊 Statistics")
    
    # Formatage du titre des statistiques
    title_cell = ws.cell(row=1, column=1, value="📊 Facial Analysis Statistics")
    title_cell.font = Font(size=16, bold=True, color="FF4A90E2")
    title_cell.alignment = Alignment(horizontal="center")
    ws.merge_cells('A1:D1')
    
    row = 3
    
    # Statistiques générales
    stats = [
        ("📁 Total Images", len(df)),
        ("✅ Faces Detected", int(df['face_detected'].sum()) if 'face_detected' in df.columns else 0),
        ("📈 Detection Rate", f"{(df['face_detected'].sum()/len(df)*100):.1f}%" if 'face_detected' in df.columns else "N/A"),
    ]
    
    for label, value in stats:
        ws.cell(row=row, column=1, value=label).font = Font(bold=True)
        ws.cell(row=row, column=2, value=value)
        row += 1
    
    row += 2
    
    # Statistiques de genre
    if 'gender' in df.columns:
        ws.cell(row=row, column=1, value="👥 Gender Distribution").font = Font(bold=True, size=12)
        row += 1
        gender_counts = df['gender'].value_counts()
        for gender, count in gender_counts.items():
            if pd.notna(gender):
                ws.cell(row=row, column=1, value=f"  {gender}")
                ws.cell(row=row, column=2, value=int(count))
                row += 1
    
    row += 2
    
    # Statistiques d'émotions
    if 'emotion' in df.columns:
        ws.cell(row=row, column=1, value="😊 Emotion Distribution").font = Font(bold=True, size=12)
        row += 1
        emotion_counts = df['emotion'].value_counts().head(8)
        for emotion, count in emotion_counts.items():
            if pd.notna(emotion):
                ws.cell(row=row, column=1, value=f"  {emotion}")
                ws.cell(row=row, column=2, value=int(count))
                row += 1
    
    row += 2
    
    # Statistiques de qualité
    if 'quality_score' in df.columns:
        ws.cell(row=row, column=1, value="⭐ Quality Statistics").font = Font(bold=True, size=12)
        row += 1
        quality_stats = [
            ("Average Quality", f"{df['quality_score'].mean():.3f}"),
            ("High Quality (≥0.7)", int((df['quality_score'] >= 0.7).sum())),
            ("Medium Quality (0.4-0.7)", int(((df['quality_score'] >= 0.4) & (df['quality_score'] < 0.7)).sum())),
            ("Low Quality (<0.4)", int((df['quality_score'] < 0.4).sum()))
        ]
        
        for label, value in quality_stats:
            ws.cell(row=row, column=1, value=f"  {label}")
            ws.cell(row=row, column=2, value=value)
            row += 1

def main():
    """Fonction principale"""
    print("🚀 Début du traitement des données...")
    
    # Vérification de l'existence du fichier
    if not os.path.exists(csv_file):
        print("❌ Fichier introuvable!")
        return
    
    # Création du nettoyeur de données
    cleaner = FacialDataCleaner(csv_file)
    
    # Chargement et nettoyage des données
    if not cleaner.load_data():
        return
    
    if not cleaner.clean_data():
        return
    
    # Sauvegarde du CSV nettoyé
    cleaned_csv = csv_file.replace(".csv", "_cleaned.csv")
    cleaner.cleaned_df.to_csv(cleaned_csv, index=False)
    print(f"💾 Données nettoyées sauvegardées: {cleaned_csv}")
    
    # Création d'Excel coloré
    excel_file = csv_file.replace(".csv", "_enhanced.xlsx")
    cleaner.cleaned_df.to_excel(excel_file, sheet_name="🎯 Facial Analysis", index=False)
    
    # Application de la mise en forme
    wb = load_workbook(excel_file)
    ws = wb.active
    
    styler = ExcelStyler()
    styler.apply_styles(ws, cleaner.cleaned_df)
    
    # Ajout de la feuille de statistiques
    create_statistics_sheet(wb, cleaner.cleaned_df)
    
    wb.save(excel_file)
    
    print(f"🎨 Fichier Excel amélioré créé: {excel_file}")
    print("✨ Fonctionnalités ajoutées:")
    print("   • Nettoyage et amélioration de la mise en forme des données")
    print("   • Scores de qualité globaux")
    print("   • Classification des âges")
    print("   • Couleurs interactives")
    print("   • Feuille de statistiques détaillée")
    print("   • Sauvegarde du CSV nettoyé")
    
    # Affichage d'un résumé rapide
    total = len(cleaner.cleaned_df)
    detected = cleaner.cleaned_df['face_detected'].sum() if 'face_detected' in cleaner.cleaned_df.columns else 0
    print(f"\n📊 Résumé rapide: {detected}/{total} visages détectés ({detected/total*100:.1f}%)")

if __name__ == "__main__":
    main()