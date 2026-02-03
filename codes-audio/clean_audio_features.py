import pandas as pd 
from openpyxl.styles import PatternFill

print("🚀 Début du nettoyage et de la coloration...")

# ✅ 1. Lecture du fichier original
df = pd.read_csv("advanced_audio_features_corrected.csv")
print(f"📂 {len(df)} lignes lues du fichier original")

# ✅ 2. Organisation des colonnes de manière logique
ordered_columns = [
    "video", "chunk", "file", "duration", "energy",
    "rms_mean", "rms_std", 
    "pauses_rms_method", "pauses_onset_method", "pauses_average", "pauses_per_minute",
    
    # Caractéristiques prosodiques
    "pitch_mean", "pitch_std", "pitch_range", "pitch_slope", "jitter",
    
    # Caractéristiques spectrales
    "spectral_centroid_mean", "spectral_centroid_std", 
    "spectral_rolloff_mean", "spectral_bandwidth_mean", 
    "zcr_mean", "spectral_contrast_1", "spectral_contrast_2", 
    "chroma_1", "chroma_2",
    
    # Caractéristiques de rythme
    "tempo", "speech_rate", "onset_density", "rhythm_regularity",
]

# Ajout des colonnes MFCCs (13 mean et 13 std)
for i in range(1, 14):
    ordered_columns.append(f"mfcc_{i}_mean")
for i in range(1, 14):
    ordered_columns.append(f"mfcc_{i}_std")

# À la fin, ajout de embedding
ordered_columns.append("embedding")

# Sélection uniquement des colonnes existantes
ordered_columns = [col for col in ordered_columns if col in df.columns]

# ✅ 3. Organisation des données
df_clean = df[ordered_columns]

# ✅ 4. Sauvegarde de la version nettoyée en CSV
df_clean.to_csv("advanced_audio_features_fixed_clean.csv", index=False)
print("✅ Créé: advanced_audio_features_fixed_clean.csv")

# ✅ 5. Création de la version colorée Excel
print("🎨 Début de la création de la version colorée...")

# Liste des couleurs
colors = ["FFDDC1", "FFABAB", "FFC3A0", "D5AAFF", "85E3FF", "B9FBC0", "FF9CEE", 
          "FFE5B4", "C4E17F", "76D7EA", "F8BBD9", "E2C2FF", "FFD93D", "A8E6CF"]

# Création d'Excel avec coloration
excel_file = "advanced_audio_features_fixed_colored.xlsx"
with pd.ExcelWriter(excel_file, engine="openpyxl") as writer:
    df_clean.to_excel(writer, index=False, sheet_name="AudioFeatures")
    ws = writer.sheets["AudioFeatures"]
    
    # Coloration de chaque colonne avec une couleur différente
    for col_idx, col in enumerate(df_clean.columns, start=1):
        fill_color = PatternFill(
            start_color=colors[(col_idx - 1) % len(colors)],
            end_color=colors[(col_idx - 1) % len(colors)],
            fill_type="solid"
        )
        for row in range(1, len(df_clean) + 2):  # +2 pour l'en-tête et les données
            ws.cell(row=row, column=col_idx).fill = fill_color

print("✅ Créé: advanced_audio_features_fixed_colored.xlsx")

# ✅ 6. Statistiques rapides
print(f"\n📊 Statistiques:")
print(f"Nombre de fichiers: {len(df_clean)}")
print(f"Nombre de colonnes: {len(df_clean.columns)}")
if "duration" in df_clean.columns:
    print(f"Durée moyenne: {df_clean['duration'].mean():.2f} secondes")

print("\n🎉 Terminé! Les deux fichiers ont été créés avec succès:")
print("   📋 advanced_audio_features_fixed_clean.csv")
print("   🎨 advanced_audio_features_fixed_colored.xlsx")